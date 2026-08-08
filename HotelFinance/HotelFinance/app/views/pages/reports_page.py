from datetime import datetime
from pathlib import Path
import calendar

from PySide6.QtCore import Qt
from PySide6.QtGui import QColor, QPainter
from PySide6.QtWidgets import ( QComboBox, QFrame, QGridLayout, QHBoxLayout, QLabel, QMessageBox, QPushButton, QVBoxLayout, QWidget, QFileDialog, )
from PySide6.QtCharts import ( QBarCategoryAxis, QBarSeries, QBarSet, QChart, QChartView, QValueAxis, )

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from app.views.widgets.stat_card import StatCard
from app.controllers.transaction_controller import TransactionController

class ReportsPage(QWidget):
    """Reports and analytics page."""

    def __init__(self, session):
        super().__init__()

        self.session = session
        self.transaction_controller = TransactionController(session)

        self.init_ui()
        self.generate_report()

    # =========================================================
    # MAIN UI
    # =========================================================

    def init_ui(self):
        main_layout = QVBoxLayout(self)

        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(18)

        self.create_header(main_layout)
        self.create_report_filters(main_layout)
        self.create_financial_overview(main_layout)
        self.create_daily_performance(main_layout)
        self.create_category_analysis(main_layout)

        main_layout.addStretch()

    # =========================================================
    # HEADER
    # =========================================================

    def create_header(self, parent_layout):
        header_layout = QHBoxLayout()
        header_layout.setSpacing(16)

        title = QLabel("Reports")
        title.setObjectName("pageTitle")

        self.current_date_label = QLabel( datetime.now().strftime("%A, %d %b %Y") )
        self.current_date_label.setObjectName("dateLabel")

        export_layout = QHBoxLayout()
        export_layout.setSpacing(10)

        self.export_pdf_button = QPushButton("Export PDF")
        self.export_pdf_button.clicked.connect(self.export_pdf)

        self.export_excel_button = QPushButton("Export Excel")
        self.export_excel_button.clicked.connect(self.export_excel)

        export_layout.addWidget(self.export_pdf_button)
        export_layout.addWidget(self.export_excel_button)

        header_layout.addWidget(title)
        header_layout.addStretch()
        header_layout.addLayout(export_layout)
        header_layout.addWidget(self.current_date_label)

        parent_layout.addLayout(header_layout)

    # =========================================================
    # REPORT FILTERS
    # =========================================================

    def create_report_filters(self, parent_layout):
        filter_frame = QFrame()
        filter_frame.setObjectName("reportFilterCard")
        filter_frame.setMinimumHeight(80)

        filter_layout = QHBoxLayout(filter_frame)

        filter_layout.setContentsMargins( 20, 15, 20, 15, )

        filter_layout.setSpacing(16)

        self.month_combo = QComboBox()

        self.month_combo.addItems( [ "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December", ] )

        self.month_combo.setFixedWidth(170)

        self.month_combo.setCurrentIndex( datetime.now().month - 1 )

        self.year_combo = QComboBox()

        current_year = datetime.now().year

        for year in range( current_year - 5, current_year + 1, ):
            self.year_combo.addItem(str(year))

        self.year_combo.setFixedWidth(120)

        self.year_combo.setCurrentText( str(current_year) )

        self.generate_button = QPushButton( "Generate Report" )

        self.generate_button.setFixedWidth(180)

        self.generate_button.clicked.connect( self.generate_report )

        filter_layout.addWidget( QLabel("Report Period:") )

        filter_layout.addWidget( self.month_combo )

        filter_layout.addWidget( self.year_combo )

        filter_layout.addStretch()

        filter_layout.addWidget( self.generate_button )

        parent_layout.addWidget( filter_frame )

    # =========================================================
    # FINANCIAL OVERVIEW
    # =========================================================

    def create_financial_overview(self, parent_layout):
        overview_title = QLabel( "Financial Overview" )

        overview_title.setObjectName( "sectionTitle" )

        parent_layout.addWidget( overview_title )

        cards_layout = QGridLayout()
        cards_layout.setSpacing(15)

        self.income_card = StatCard( "Monthly Income", "₹0" )

        self.expense_card = StatCard( "Monthly Expense", "₹0" )

        self.profit_card = StatCard( "Net Profit", "₹0" )

        self.transaction_card = StatCard( "Transactions", "0" )

        cards_layout.addWidget( self.income_card, 0, 0 )

        cards_layout.addWidget( self.expense_card, 0, 1 )

        cards_layout.addWidget( self.profit_card, 0, 2 )

        cards_layout.addWidget( self.transaction_card, 0, 3 )

        for column in range(4):
            cards_layout.setColumnStretch( column, 1 )

        parent_layout.addLayout( cards_layout )

    # =========================================================
    # DAILY / WEEKLY PERFORMANCE
    # =========================================================

    def create_daily_performance(self, parent_layout):
        performance_header = QHBoxLayout()

        performance_title = QLabel( "Weekly Performance" )

        performance_title.setObjectName( "sectionTitle" )

        self.week_combo = QComboBox()

        self.week_combo.setFixedWidth(140)

        self.week_combo.currentIndexChanged.connect( self.update_selected_week_chart )

        performance_header.addWidget( performance_title )

        performance_header.addStretch()

        performance_header.addWidget( QLabel("View:") )

        performance_header.addWidget( self.week_combo )

        parent_layout.addLayout( performance_header )

        self.daily_performance_chart = ( self.create_chart_placeholder( "Income vs Expense" ) )

        self.daily_performance_chart.setMinimumHeight( 380 )

        parent_layout.addWidget( self.daily_performance_chart )

    # =========================================================
    # CATEGORY ANALYSIS
    # =========================================================

    def create_category_analysis(self, parent_layout):
        category_title = QLabel( "Category Analysis" )

        category_title.setObjectName( "sectionTitle" )

        parent_layout.addWidget( category_title )

        charts_layout = QGridLayout()
        charts_layout.setSpacing(15)

        self.income_breakdown_chart = ( self.create_chart_placeholder( "Income Breakdown" ) )

        self.expense_breakdown_chart = ( self.create_chart_placeholder( "Expense Breakdown" ) )

        charts_layout.addWidget( self.income_breakdown_chart, 0, 0 )

        charts_layout.addWidget( self.expense_breakdown_chart, 0, 1 )

        charts_layout.setColumnStretch( 0, 1 )

        charts_layout.setColumnStretch( 1, 1 )

        parent_layout.addLayout( charts_layout )

    # =========================================================
    # CHART PLACEHOLDER
    # =========================================================

    def create_chart_placeholder(self, title):
        frame = QFrame()

        frame.setObjectName( "chartCard" )

        frame.setMinimumHeight( 360 )

        layout = QVBoxLayout(frame)

        layout.setContentsMargins( 16, 16, 16, 16 )

        layout.setSpacing(10)

        title_label = QLabel(title)

        title_label.setObjectName( "chartTitle" )

        title_label.setAlignment( Qt.AlignLeft )

        layout.addWidget( title_label )

        chart_container = QWidget()

        chart_layout = QVBoxLayout( chart_container )

        chart_layout.setContentsMargins( 0, 0, 0, 0 )

        chart_layout.setSpacing(0)

        layout.addWidget( chart_container, 1 )

        frame.chart_container = ( chart_container )

        frame.setContentsMargins( 0, 0, 0, 0 )

        return frame

    # =========================================================
    # GENERATE REPORT
    # =========================================================

    def generate_report(self):
        month = ( self.month_combo.currentIndex() + 1 )

        year = int( self.year_combo.currentText() )

        income = ( self.transaction_controller .get_monthly_income( month, year ) )

        expense = ( self.transaction_controller .get_monthly_expense( month, year ) )

        transactions = ( self.transaction_controller .get_monthly_transaction_count( month, year ) )

        profit = income - expense

        self.income_card.set_value( f"₹ {income:.2f}" )

        self.expense_card.set_value( f"₹ {expense:.2f}" )

        self.profit_card.set_value( f"₹ {profit:.2f}" )

        self.transaction_card.set_value( str(transactions) )

        # -----------------------------------------------------
        # Daily data for selected month
        # -----------------------------------------------------

        self.daily_data = ( self.transaction_controller .get_daily_income_expense( month, year ) )

        # -----------------------------------------------------
        # Update week selector
        # -----------------------------------------------------

        self.update_week_selector( month, year )

        # -----------------------------------------------------
        # Category data
        # -----------------------------------------------------

        income_categories = ( self.transaction_controller .get_income_by_category( month, year ) )

        expense_categories = ( self.transaction_controller .get_expense_by_category( month, year ) )

        self.show_category_chart( self.income_breakdown_chart, income_categories )

        self.show_category_chart( self.expense_breakdown_chart, expense_categories )

    # =========================================================
    # WEEK SELECTOR
    # =========================================================

    def update_week_selector( self, month, year ):
        days_in_month = calendar.monthrange( year, month )[1]

        week_count = ( (days_in_month + 6) // 7 )

        self.week_combo.blockSignals( True )

        self.week_combo.clear()

        for week in range( 1, week_count + 1 ):
            start_day = ( (week - 1) * 7 ) + 1

            end_day = min( week * 7, days_in_month )

            self.week_combo.addItem( f"Week {week}" )

        self.week_combo.blockSignals( False )

        self.week_combo.setCurrentIndex( 0 )

        self.update_selected_week_chart()

    # =========================================================
    # UPDATE SELECTED WEEK
    # =========================================================

    def update_selected_week_chart(self):
        if not hasattr( self, "daily_data" ):
            return

        if not self.daily_data:
            self.show_weekly_bar_chart( [] )
            return

        week_number = ( self.week_combo.currentIndex() + 1 )

        start_day = ( (week_number - 1) * 7 ) + 1

        end_day = ( week_number * 7 )

        selected_week_data = []

        for row in self.daily_data:
            day = int(row[0])

            if ( start_day <= day <= end_day ):
                selected_week_data.append( row )

        self.show_weekly_bar_chart( selected_week_data )

    # =========================================================
    # WEEKLY BAR CHART
    # =========================================================

    def show_weekly_bar_chart( self, weekly_data ):
        layout = ( self.daily_performance_chart .chart_container .layout() )

        # Clear old chart
        while layout.count():
            item = layout.takeAt(0)

            widget = item.widget()

            if widget:
                widget.deleteLater()

        if not weekly_data:
            empty_label = QLabel( "No transaction data available for this week." )

            empty_label.setAlignment( Qt.AlignCenter )

            empty_label.setObjectName( "emptyChartLabel" )

            layout.addWidget( empty_label )

            return

        # -----------------------------------------------------
        # Income and Expense bar sets
        # -----------------------------------------------------

        income_set = QBarSet( "Income" )

        expense_set = QBarSet( "Expense" )

        income_set.setColor( QColor("#2196F3") )

        expense_set.setColor( QColor("#FF9800") )

        categories = []

        maximum = 0

        for day, income, expense in weekly_data:

            day_number = int(day)

            income_value = float( income )

            expense_value = float( expense )

            # Actual date shown on X-axis
            month = ( self.month_combo.currentIndex() + 1 )

            year = int( self.year_combo.currentText() )

            date_value = datetime( year, month, day_number )

            date_label = date_value.strftime( "%d %b" )

            categories.append( date_label )

            income_set.append( income_value )

            expense_set.append( expense_value )

            maximum = max( maximum, income_value, expense_value )

        # -----------------------------------------------------
        # Bar series
        # -----------------------------------------------------

        series = QBarSeries()

        series.append( income_set )

        series.append( expense_set )

        # Slightly narrower bars for cleaner spacing
        series.setBarWidth( 0.65 )

        # -----------------------------------------------------
        # Chart
        # -----------------------------------------------------

        chart = QChart()

        chart.addSeries( series )

        chart.setTitle( "" )

        chart.setAnimationOptions( QChart.SeriesAnimations )

        chart.legend().setVisible( True )

        chart.legend().setAlignment( Qt.AlignBottom )

        chart.layout().setContentsMargins( 0, 0, 0, 0 )

        chart.setBackgroundRoundness( 0 )

        # -----------------------------------------------------
        # X Axis - Actual Dates
        # -----------------------------------------------------

        axis_x = QBarCategoryAxis()

        axis_x.append( categories )

        axis_x.setLabelsAngle( 0 )

        axis_x.setGridLineVisible( False )

        chart.addAxis( axis_x, Qt.AlignBottom )

        series.attachAxis( axis_x )

        # -----------------------------------------------------
        # Y Axis
        # -----------------------------------------------------

        axis_y = QValueAxis()

        if maximum > 0:
            axis_maximum = ( maximum * 1.20 )
        else:
            axis_maximum = 100

        axis_y.setRange( 0, axis_maximum )

        axis_y.setTickCount( 6 )

        # Use Rs instead of ₹ here.
        # This avoids the '?' problem that Qt Charts
        # can show with some fonts.
        axis_y.setLabelFormat( "Rs %.0f" )

        axis_y.setTitleText( "Amount" )

        axis_y.setGridLineVisible( True )

        chart.addAxis( axis_y, Qt.AlignLeft )

        series.attachAxis( axis_y )

        # -----------------------------------------------------
        # Chart View
        # -----------------------------------------------------

        chart_view = QChartView( chart )

        chart_view.setRenderHint( QPainter.Antialiasing )

        chart_view.setMinimumHeight( 300 )

        layout.addWidget( chart_view )

    # =========================================================
    # CATEGORY BAR CHARTS
    # =========================================================

    def show_category_chart( self, chart_frame, data ):
        layout = ( chart_frame .chart_container .layout() )

        # Clear old chart
        while layout.count():
            item = layout.takeAt(0)

            widget = item.widget()

            if widget:
                widget.deleteLater()

        if not data:
            empty_label = QLabel( "No data available for this period." )

            empty_label.setAlignment( Qt.AlignCenter )

            empty_label.setObjectName( "emptyChartLabel" )

            layout.addWidget( empty_label )

            return

        series = QBarSeries()

        categories = []

        colors = [ QColor("#4CAF50"), QColor("#2196F3"), QColor("#FF9800"), QColor("#9C27B0"), QColor("#F44336"), QColor("#00BCD4"), QColor("#795548"), QColor("#607D8B"), ]

        maximum = 0

        for index, ( name, amount ) in enumerate(data):

            bar = QBarSet( name )

            value = float( amount )

            bar.append( value )

            bar.setColor( colors[ index % len(colors) ] )

            series.append( bar )

            categories.append( name )

            maximum = max( maximum, value )

        chart = QChart()

        chart.addSeries( series )

        chart.setTitle( "" )

        chart.legend().setVisible( True )

        chart.legend().setAlignment( Qt.AlignBottom )

        chart.setAnimationOptions( QChart.SeriesAnimations )

        chart.layout().setContentsMargins( 0, 0, 0, 0 )

        chart.setBackgroundRoundness( 0 )

        axis_x = QBarCategoryAxis()

        axis_x.append( categories )

        axis_x.setLabelsAngle( 0 )

        axis_x.setGridLineVisible(False)

        chart.addAxis(axis_x,Qt.AlignBottom)

        series.attachAxis(axis_x)

        axis_y = QValueAxis()

        axis_y.setRange(0, maximum * 1.2 if maximum else 100)

        axis_y.setTickCount( 5 )

        axis_y.setLabelFormat( "Rs %.0f" )

        axis_y.setGridLineVisible( True )

        chart.addAxis( axis_y, Qt.AlignLeft )

        series.attachAxis( axis_y )

        chart_view = QChartView( chart )

        chart_view.setRenderHint( QPainter.Antialiasing )

        layout.addWidget( chart_view )

    # =========================================================
    # EXPORT PDF
    # =========================================================

    def export_pdf(self):
        month = ( self.month_combo.currentIndex() + 1 )

        year = int( self.year_combo.currentText() )

        month_name = ( self.month_combo.currentText() )

        default_name = ( f"Hotel_Report_" f"{month_name}_" f"{year}.pdf" )

        file_path, _ = ( QFileDialog.getSaveFileName( self, "Save PDF Report", default_name, "PDF Files (*.pdf)", ) )

        if not file_path:
            return

        if Path(file_path).exists():
            confirm = QMessageBox.question( self, "Overwrite File?", ( f"The file " f"'{Path(file_path).name}' " f"already exists. " f"Do you want to replace it?" ), QMessageBox.Yes | QMessageBox.No, )

            if confirm != QMessageBox.Yes:
                return

        income = ( self.transaction_controller .get_monthly_income( month, year ) )

        expense = ( self.transaction_controller .get_monthly_expense( month, year ) )

        profit = income - expense

        transactions = ( self.transaction_controller .get_monthly_transaction_count( month, year ) )

        daily_data = ( self.transaction_controller .get_daily_income_expense( month, year ) )

        income_categories = ( self.transaction_controller .get_income_by_category( month, year ) )

        expense_categories = ( self.transaction_controller .get_expense_by_category( month, year ) )

        try:
            self._create_pdf( file_path, month_name, year, income, expense, profit, transactions, daily_data, income_categories, expense_categories, )

            QMessageBox.information( self, "Report Saved", "PDF report was saved successfully.", )

        except Exception as exc:
            QMessageBox.critical( self, "Export Failed", ( "Unable to create " f"PDF report.\n{exc}" ), )

    # =========================================================
    # CREATE PDF
    # =========================================================

    def _create_pdf( self, file_path, month_name, year, income, expense, profit, transactions, daily_data, income_categories, expense_categories, ):
        pdf = canvas.Canvas( file_path, pagesize=A4 )

        width, height = A4

        margin = 24 * mm

        y = height - margin

        pdf.setFont( "Helvetica-Bold", 18 )

        pdf.drawString( margin, y, "HOTEL EXPENSE TRACKER" )

        y -= 18

        pdf.setFont( "Helvetica", 12 )

        pdf.drawString( margin, y, "Monthly Financial Report" )

        pdf.drawRightString( width - margin, y, datetime.now().strftime( "%d %b %Y" ) )

        y -= 26

        pdf.setFont( "Helvetica-Bold", 11 )

        pdf.drawString( margin, y, f"Report Period: {month_name} {year}" )

        y -= 20

        pdf.setFont( "Helvetica-Bold", 12 )

        pdf.drawString( margin, y, "Financial Overview" )

        y -= 18

        pdf.setFont( "Helvetica", 10 )

        overview_lines = [ ( "Monthly Income:", f"₹ {income:.2f}" ), ( "Monthly Expense:", f"₹ {expense:.2f}" ), ( "Net Profit:", f"₹ {profit:.2f}" ), ( "Transaction Count:", str(transactions) ), ]

        for label, value in overview_lines:
            pdf.drawString( margin, y, label )

            pdf.drawRightString( width - margin, y, value )

            y -= 16

        y -= 8

        pdf.setFont( "Helvetica-Bold", 12 )

        pdf.drawString( margin, y, "Daily Performance" )

        y -= 18

        if not daily_data:

            pdf.drawString( margin, y, "No daily transaction data available for this period." )

            y -= 18

        else:

            pdf.setFont( "Helvetica-Bold", 10 )

            pdf.drawString( margin, y, "Day" )

            pdf.drawString( margin + 80, y, "Income" )

            pdf.drawString( margin + 180, y, "Expense" )

            pdf.drawRightString( width - margin, y, "Net" )

            y -= 14

            pdf.setFont( "Helvetica", 10 )

            for ( day, income_value, expense_value ) in daily_data:

                if y < margin + 40:

                    pdf.showPage()

                    y = height - margin

                    pdf.setFont( "Helvetica-Bold", 10 )

                    pdf.drawString( margin, y, "Day" )

                    pdf.drawString( margin + 80, y, "Income" )

                    pdf.drawString( margin + 180, y, "Expense" )

                    pdf.drawRightString( width - margin, y, "Net" )

                    y -= 14

                    pdf.setFont( "Helvetica", 10 )

                net = ( income_value - expense_value )

                pdf.drawString( margin, y, str(day) )

                pdf.drawString( margin + 80, y, f"₹ {income_value:.2f}" )

                pdf.drawString( margin + 180, y, f"₹ {expense_value:.2f}" )

                pdf.drawRightString( width - margin, y, f"₹ {net:.2f}" )

                y -= 14

        y -= 12

        pdf.setFont( "Helvetica-Bold", 12 )

        pdf.drawString( margin, y, "Income by Category" )

        y -= 18

        if not income_categories:

            pdf.drawString( margin, y, "No income categories available for this period." )

            y -= 18

        else:

            pdf.setFont( "Helvetica-Bold", 10 )

            pdf.drawString( margin, y, "Category" )

            pdf.drawRightString( width - margin, y, "Amount" )

            y -= 14

            pdf.setFont( "Helvetica", 10 )

            for category, amount in income_categories:

                if y < margin + 40:

                    pdf.showPage()

                    y = height - margin

                pdf.drawString( margin, y, str(category) )

                pdf.drawRightString( width - margin, y, f"₹ {float(amount):.2f}" )

                y -= 14

        y -= 12

        pdf.setFont( "Helvetica-Bold", 12 )

        pdf.drawString( margin, y, "Expense by Category" )

        y -= 18

        if not expense_categories:

            pdf.drawString( margin, y, "No expense categories available for this period." )

            y -= 18

        else:

            pdf.setFont( "Helvetica-Bold", 10 )

            pdf.drawString( margin, y, "Category" )

            pdf.drawRightString( width - margin, y, "Amount" )

            y -= 14

            pdf.setFont( "Helvetica", 10 )

            for category, amount in expense_categories:

                if y < margin + 40:

                    pdf.showPage()

                    y = height - margin

                pdf.drawString( margin, y, str(category) )

                pdf.drawRightString( width - margin, y, f"₹ {float(amount):.2f}" )

                y -= 14

        pdf.save()

    # =========================================================
    # EXPORT EXCEL
    # =========================================================

    def export_excel(self):
        month = ( self.month_combo.currentIndex() + 1 )

        year = int( self.year_combo.currentText() )

        month_name = ( self.month_combo.currentText() )

        default_name = ( f"Hotel_Report_" f"{month_name}_" f"{year}.xlsx" )

        file_path, _ = ( QFileDialog.getSaveFileName( self, "Save Excel Report", default_name, "Excel Files (*.xlsx)", ) )

        if not file_path:
            return

        if Path(file_path).exists():

            confirm = QMessageBox.question( self, "Overwrite File?", ( f"The file " f"'{Path(file_path).name}' " f"already exists. " f"Do you want to replace it?" ), QMessageBox.Yes | QMessageBox.No, )

            if confirm != QMessageBox.Yes:
                return

        income = ( self.transaction_controller .get_monthly_income( month, year ) )

        expense = ( self.transaction_controller .get_monthly_expense( month, year ) )

        profit = income - expense

        transactions = ( self.transaction_controller .get_monthly_transaction_count( month, year ) )

        daily_data = ( self.transaction_controller .get_daily_income_expense( month, year ) )

        income_categories = ( self.transaction_controller .get_income_by_category( month, year ) )

        expense_categories = ( self.transaction_controller .get_expense_by_category( month, year ) )

        try:

            self._create_excel( file_path, month_name, year, income, expense, profit, transactions, daily_data, income_categories, expense_categories, )

            QMessageBox.information( self, "Report Saved", "Excel report was saved successfully.", )

        except Exception as exc:

            QMessageBox.critical( self, "Export Failed", ( "Unable to create " f"Excel report.\n{exc}" ), )

    # =========================================================
    # CREATE EXCEL
    # =========================================================

    def _create_excel( self, file_path, month_name, year, income, expense, profit, transactions, daily_data, income_categories, expense_categories, ):
        wb = Workbook()

        summary = wb.active

        summary.title = "Summary"

        bold = Font( bold=True )

        border = Border( bottom=Side( style="thin" ) )

        alignment = Alignment( horizontal="left", vertical="center" )

        rows = [
            ( "Report Month", month_name ),
            ( "Report Year", year ),
            ( "Monthly Income", float(income) ),
            ( "Monthly Expense", float(expense) ),
            ( "Net Profit", float(profit) ),
            ( "Transaction Count", transactions ),
        ]

        for row_index, ( label, value ) in enumerate( rows, start=1 ):

            summary.cell( row=row_index, column=1, value=label ).font = bold

            summary.cell( row=row_index, column=2, value=value )

        summary.column_dimensions[ "A" ].width = 26

        summary.column_dimensions[ "B" ].width = 20

        # -----------------------------------------------------
        # Daily Performance Sheet
        # -----------------------------------------------------

        daily_sheet = wb.create_sheet( "Daily Performance" )

        headers = [ "Day", "Income", "Expense", "Net" ]

        for idx, header in enumerate( headers, start=1 ):

            cell = daily_sheet.cell( row=1, column=idx, value=header )

            cell.font = bold
            cell.alignment = alignment
            cell.border = border

        for row_index, ( day, income_value, expense_value ) in enumerate( daily_data, start=2 ):

            net_value = float( income_value - expense_value )

            daily_sheet.cell( row=row_index, column=1, value=int(day) )

            daily_sheet.cell( row=row_index, column=2, value=float( income_value ) )

            daily_sheet.cell( row=row_index, column=3, value=float( expense_value ) )

            daily_sheet.cell( row=row_index, column=4, value=net_value )

        for col in [ "B", "C", "D" ]:

            daily_sheet.column_dimensions[ col ].width = 18

            for row in range( 2, len(daily_data) + 2 ):

                daily_sheet[ f"{col}{row}" ].number_format = ( '₹#,##0.00' )

        daily_sheet.column_dimensions[ "A" ].width = 10

        daily_sheet.freeze_panes = "A2"

        # -----------------------------------------------------
        # Category Sheets
        # -----------------------------------------------------

        income_sheet = wb.create_sheet( "Income by Category" )

        expense_sheet = wb.create_sheet( "Expense by Category" )

        for sheet, data in [ ( income_sheet, income_categories ), ( expense_sheet, expense_categories ), ]:

            sheet.cell( row=1, column=1, value="Category" ).font = bold

            sheet.cell( row=1, column=2, value="Amount" ).font = bold

            sheet.column_dimensions[ "A" ].width = 28

            sheet.column_dimensions[ "B" ].width = 18

            sheet.freeze_panes = "A2"

            for row_index, ( category, amount ) in enumerate( data, start=2 ):

                sheet.cell( row=row_index, column=1, value=str(category) )

                sheet.cell( row=row_index, column=2, value=float(amount) )

                sheet.cell( row=row_index, column=2 ).number_format = ( '₹#,##0.00' )

        wb.save( file_path)